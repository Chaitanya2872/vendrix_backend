from pathlib import Path
from io import BytesIO
from sqlalchemy import select
from app.core.config import settings
from app.db.session import SessionLocal
from app.models import Document
from app.workers.celery_app import celery_app
from app.workers.vision import decode_image, read_text


def extract_document_fields(document_type: str, text: str) -> dict:
    """Deterministic first-pass extraction; replace/extend with domain ML rules."""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return {"document_type": document_type, "text": text[:10000], "preview_lines": lines[:20]}


def process_vendor_document_now(document_id: str) -> dict:
    with SessionLocal() as db:
        document = db.get(Document, document_id)
        if not document: return {"document_id": document_id, "status": "not_found"}
        path = Path(settings.storage_path) / document.object_key
        if not path.exists(): raise FileNotFoundError(path)
        text = ""
        extension = path.suffix.lower()
        if extension == ".pdf":
            import fitz
            with fitz.open(path) as pdf:
                text = "\n".join(page.get_text() for page in pdf)
                if not text.strip():
                    pages = []
                    for page in pdf:
                        pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                        pages.append(read_text(decode_image(pixmap.tobytes("png"))))
                    text = "\n".join(pages)
        elif extension in {".jpg", ".jpeg", ".png", ".webp"}:
            text = read_text(decode_image(path.read_bytes()))
        elif extension == ".docx":
            from docx import Document as WordDocument
            word_document = WordDocument(path)
            text = "\n".join(paragraph.text for paragraph in word_document.paragraphs)
        elif extension == ".xlsx":
            from openpyxl import load_workbook
            workbook = load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)
            text = "\n".join(" | ".join(str(value) for value in row if value is not None) for sheet in workbook.worksheets for row in sheet.iter_rows(values_only=True))
        else:
            raise ValueError(f"Unsupported document extension: {extension}")
        fields = extract_document_fields(document.document_type, text)
        document.extracted_fields, document.status = fields, "REVIEW_REQUIRED"; db.commit()
        return {"document_id": document_id, "status": "review_required", "fields": fields}


@celery_app.task(bind=True, autoretry_for=(Exception,), retry_backoff=True, max_retries=3)
def process_vendor_document(self, document_id: str) -> dict:
    return process_vendor_document_now(document_id)
